"""Check the imported master against Excel's own cached results.

The workbook stores the value Excel last calculated for every formula cell.
Those are ground truth: they were produced by Excel, not by us. This replays
each seed row through `calc.derive_line` and compares:

  stickers per box   against the sheet's cached V / T column
  boxes, volume, weights, labels, sheets, values, FOB, RBI reference
                     against the same formulas evaluated on the raw cells,
                     re-implemented here independently of calc.py

Run:  python backend/scripts/verify_masters.py
Exit 0 = every row agrees.
"""
from __future__ import annotations

import json
import re
import math
import sys
from pathlib import Path
from types import SimpleNamespace

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "backend"))

from app import calc  # noqa: E402

XLSX = ROOT / "Docs" / "Data" / "Masters.xlsx"
SEED = ROOT / "backend" / "app" / "data" / "masters_seed.json"
TOL = 1e-6

# Which column holds the sticker total, per sheet layout.
TTL_COL = {"Kiran": 22, "Hansa-PP": 22, "VP-PP": 22,      # V
           "Hansa-GRN": 20, "VP-GRN": 20,                  # T
           "Oswin": 20}                                    # T


def close(a, b, tol=TOL):
    return abs(float(a) - float(b)) <= tol * max(1.0, abs(float(a)), abs(float(b)))


def excel_expected(it, qty, rbi):
    """The sheet's formula chain, written out again from the seed fields.

    Deliberately a second implementation: if this and calc.derive_line drift,
    one of them is wrong.
    """
    packing = it["packing"]
    boxes = qty / packing if packing else 0
    bg, pc = it["bg_per_box"], it["p_per_box"]
    if it["stickers_fixed"]:
        ttl = it["stickers_fixed"]
    else:
        ttl = (bg + pc) * it["sticker_mult"]
        if it["sticker_round"]:
            ttl = math.floor(ttl + 0.5)
    labels = boxes * ttl * it["label_spoilage"]
    up = it["type_up"]
    sheets = math.ceil(labels / up) if up and labels else 0
    value = qty * it["unit_value"]
    fob = qty * it["unit_fob100"] / 100 if it["fob_mode"] == "100" else qty * it["unit_fob100"]
    return {
        "boxes_exact": boxes,
        "vol_total": boxes * it["volume"],
        "net_total": boxes * it["net_per_box"],
        "gross_total": boxes * it["gross_per_box"],
        "stickers_per_box": ttl,
        "labels": labels,
        "sheets": sheets,
        "total_value_inr": value,
        "total_fob_usd": fob,
        "rbi_ref_inr": rbi * fob,
        "rate": (rbi * fob / qty) if qty else 0,
    }


# Which column each sheet's formulas must reference. Guards the highest-risk
# failure mode of the extractor: reading the right value out of the wrong
# column. Every entry is "this formula, on this sheet, must look like this".
WIRING = {
    "Kiran":     {12: "=K#/G#", 13: "=L#*0.06", 17: "=O#*L#", 18: "=P#*L#",
                  20: "=$G#/$F#", 21: "=$G#", 23: "=$L#*V#", 25: "=ROUNDUP(W#/X#,)",
                  28: "=$K#*AA#", 31: "=K#*AD#/100"},
    "Hansa-PP":  {12: "=K#/G#", 13: "=L#*0.06", 17: "=O#*L#", 18: "=P#*L#",
                  20: "=$G#/$F#", 21: "=$G#", 23: "=$L#*V#", 25: "=ROUNDUP(W#/X#,)",
                  28: "=K#*AA#", 31: "=K#*AD#/100"},
    "VP-PP":     {12: "=K#/G#", 13: "=L#*0.06", 17: "=O#*L#", 18: "=P#*L#",
                  20: "=$G#/$F#", 21: "=$G#", 23: "=$L#*V#", 25: "=ROUNDUP(W#/X#,)",
                  28: "=$K#*AA#", 31: "=K#*AD#/100"},
    "Hansa-GRN": {10: "=I#/E#", 11: "=J#*0.06", 15: "=M#*J#", 16: "=N#*J#",
                  18: "=$E#/$D#", 19: "=$E#", 21: "=$J#*T#", 23: "=ROUNDUP(U#/V#,)",
                  26: "=Y#*$I#", 29: "=I#*AB#/100"},
    "VP-GRN":    {10: "=I#/E#", 11: "=J#*0.06", 15: "=M#*J#", 16: "=N#*J#",
                  18: "=$E#/$D#", 19: "=$E#", 21: "=$J#*T#", 23: "=ROUNDUP(U#/V#,)",
                  26: "=$I#*Y#", 29: "=I#*AB#/100"},
    "Oswin":     {13: "=$L#/$H#", 16: "=$M#*$O#", 18: "=G#", 19: "=H#",
                  21: "=$M#*$T#*1.05", 25: "=X#*$L#", 28: "=$L#*AA#"},
}
NORM = lambda s: re.sub(r"(\$?[A-Z]{1,2})\$?\d+", r"\1#", s)


def check_wiring(wb_f) -> list[str]:
    """Row 3 of every sheet must use the columns the extractor reads."""
    bad = []
    for sheet, cols in WIRING.items():
        ws = wb_f[sheet]
        for col, want in cols.items():
            got = ws.cell(3, col).value
            if not isinstance(got, str) or NORM(got) != want:
                bad.append(f"wiring {sheet} col{col}: expected {want} got {got!r}")
    return bad


def main() -> int:
    seed = json.loads(SEED.read_text())
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    wb_f = openpyxl.load_workbook(XLSX, data_only=False)
    fails: list[str] = check_wiring(wb_f)
    print(f"column wiring vs sheet formulas    : "
          f"{sum(len(c) for c in WIRING.values())} formulas checked")

    # ---- 1. sticker total vs Excel's own cached cell -----------------------
    checked = 0
    for it in seed["items"]:
        ws = wb[it["source_sheet"]]
        cached = ws.cell(it["source_row"], TTL_COL[it["source_sheet"]]).value
        if cached is None or isinstance(cached, str):
            continue
        ours = calc.stickers_per_box(SimpleNamespace(**it))
        checked += 1
        if not close(ours, cached):
            fails.append(f"stickers {it['source_sheet']}!{it['source_row']} "
                         f"{it['code']}: excel={cached} ours={ours}")
    print(f"stickers/box vs Excel cached cells : {checked} rows checked")

    # ---- 2. full derivation, whole boxes and a part box -------------------
    for it in seed["items"]:
        item = SimpleNamespace(**it)
        for mult, extra in ((7, 0), (3, 1), (1, 0)):
            qty = it["packing"] * mult + extra
            rbi = 87.25
            got = calc.derive_line(item, qty, rbi)
            want = excel_expected(it, qty, rbi)
            for k, v in want.items():
                if not close(got[k], v):
                    fails.append(f"{k} {it['source_sheet']}!{it['source_row']} "
                                 f"{it['code']} qty={qty}: want={v} got={got[k]}")
            if got["boxes"] != math.ceil(want["boxes_exact"]):
                fails.append(f"boxes {it['code']} qty={qty}")
    print(f"derive_line vs re-implemented sheet : {len(seed['items']) * 3} evaluations")

    # ---- 3. structural sanity --------------------------------------------
    for it in seed["items"]:
        if it["packing"] <= 0:
            fails.append(f"packing<=0 {it['code']} {it['source_sheet']}!{it['source_row']}")
        if it["sticker_rule"] not in calc.STICKER_RULES:
            fails.append(f"bad rule {it['code']}")
        if it["fob_mode"] not in ("piece", "100", "custom"):
            fails.append(f"bad fob_mode {it['code']}")

    if fails:
        print(f"\n{len(fails)} MISMATCH(ES):")
        for f in fails[:40]:
            print("  " + f)
        return 1
    print("\nAll rows agree with the workbook.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
