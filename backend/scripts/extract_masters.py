"""Extract the client's master workbook into the seed JSON the API imports.

Source: Docs/Data/Masters.xlsx
  Sheet1                    parties — buyers, suppliers, transporters
  Kiran / Hansa-PP / VP-PP  "PP" layout   (2A master, moulded risers & fittings)
  Hansa-GRN / VP-GRN        "GRN" layout  (2A master, green range)
  Oswin                     "OSWIN" layout (7A moulded order master + tube sections)

The layouts differ in column order and in how barcode stickers are counted.
Rather than assume a rule per sheet, every row's own sticker formula is read
and reduced to three numbers, because the workbook has per-row exceptions
(VP-PP rows 9-11 use the GRN rounding, rows 11/16 multiply by 1 not 1.1,
Oswin row 121 hard-codes its total):

  stickers/box = stickers_fixed                    when the cell is a literal
               = ROUND((bg + p) * sticker_mult)    when the formula rounds
               = (bg + p) * sticker_mult           otherwise
  labels       = boxes * stickers/box * label_spoilage      (1.05 on Oswin)
  sheets       = ROUNDUP(labels / type_up)

Run:  python backend/scripts/extract_masters.py
Out:  backend/app/data/masters_seed.json
"""
from __future__ import annotations

import json
import math
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "Docs" / "Data" / "Masters.xlsx"
OUT = ROOT / "backend" / "app" / "data" / "masters_seed.json"

# Volume per box is a flat 0.06 m³ on the PP and GRN sheets (=L*0.06); the
# Oswin sheet carries a real per-item figure in its own column.
FLAT_VOLUME = 0.06


# ----------------------------------------------------------------- helpers
def _txt(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _numf(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _num_or_none(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(",", ""))
    except ValueError:
        return None


def hsn8(v) -> str:
    """3917.4 → "39174000".  The sheet stores HSN as a number formatted 0.0000,
    i.e. chapter heading + four sub-digits, so pad the fraction to four."""
    if v is None or v == "":
        return ""
    if isinstance(v, str) and not re.fullmatch(r"[\d.]+", v.strip()):
        return v.strip()
    s = f"{float(v):.4f}"            # 3917.4 -> "3917.4000"
    head, frac = s.split(".")
    return f"{head}{frac}"


def sticker_spec(sh, row, ttl_col, labels_col, sheets_col, up_col):
    """Reduce one row's sticker / label / sheet formulas to plain numbers.

    Handles every shape present in the workbook:
      =(U+T)*1.1 · =(U+T)*1 · =ROUND(SUM(T:U)*1.1,) · =S+R · a typed constant
      =$L*V · =$J*T · =$M*$T*1.05
      =ROUNDUP(W/X,) → labels per sheet from column X
      =ROUNDUP($U/125,) → the divisor is written into the formula
    """
    ttl = sh.raw(row, ttl_col)
    mult, rounds, fixed = 1.1, False, 0.0
    if isinstance(ttl, str) and ttl.startswith("="):
        m = re.search(r"\*\s*([\d.]+)\s*,?\s*\)?\s*$", ttl)
        mult = float(m.group(1)) if m else 1.0
        rounds = "ROUND(" in ttl.upper()
    elif ttl is not None:
        fixed = _numf(ttl)

    lab = sh.raw(row, labels_col)
    m = re.search(r"\*\s*([\d.]+)\s*$", lab) if isinstance(lab, str) else None
    spoilage = float(m.group(1)) if m else 1.0

    shf = sh.raw(row, sheets_col)
    if isinstance(shf, str) and (m := re.search(r"/\s*(\d+)\s*[,)]", shf)):
        type_up = int(m.group(1))          # divisor written into the formula
    else:
        type_up = int(_numf(sh.raw(row, up_col))) if up_col else 0

    return {"sticker_mult": mult, "sticker_round": rounds,
            "stickers_fixed": fixed, "label_spoilage": spoilage, "type_up": type_up}


class Sheet:
    """Cell access that prefers Excel's cached result for formula cells."""

    def __init__(self, wb_f, wb_v, name):
        self.f = wb_f[name]
        self.v = wb_v[name]
        self.name = name
        self.max_row = self.f.max_row

    def raw(self, row, col):
        return self.f.cell(row, col).value

    def val(self, row, col):
        """Cached value for a formula cell, else the literal."""
        raw = self.f.cell(row, col).value
        if isinstance(raw, str) and raw.startswith("="):
            cached = self.v.cell(row, col).value
            if isinstance(cached, str) and cached.startswith("#"):
                return None          # #DIV/0!, #REF! — treat as blank
            return cached
        return raw


# ----------------------------------------------------------------- Sheet1
def parse_parties(sh: Sheet):
    """Sheet1 holds three stacked blocks: buyers, suppliers, transporters."""
    buyers, suppliers, transports = [], [], []
    block = None
    for r in range(1, sh.max_row + 1):
        a = _txt(sh.raw(r, 1))
        low = a.lower()
        if low.startswith("buyers data"):
            block = "buyer"; continue
        if low.startswith("suppliers data"):
            block = "supplier"; continue
        if low.startswith("transporters"):
            block = "transport"; continue
        if not a:
            continue
        # header rows inside a block
        if block == "buyer" and low == "buyer name":
            continue
        if block == "supplier" and low == "code":
            continue
        if block == "transport" and low == "transport name":
            continue

        if block == "buyer":
            trading = _txt(sh.raw(r, 2))
            brand = re.sub(r"^t\s*/\s*a\s+", "", trading, flags=re.I).strip()
            buyers.append({
                "name": a,
                "brand": brand or trading,
                "addr": _txt(sh.raw(r, 3)),
                "country": _txt(sh.raw(r, 4)),
                "curr": _txt(sh.raw(r, 5)) or "USD",
                "ship_to": _txt(sh.raw(r, 6)),
                "order_no": "",
            })
        elif block == "supplier":
            suppliers.append({
                "code": a,
                "name": _txt(sh.raw(r, 2)),
                "gstin": _txt(sh.raw(r, 3)),
                "addr": _txt(sh.raw(r, 4)),
                "pin": _txt(sh.raw(r, 5)),
                "place": _txt(sh.raw(r, 6)),
                "state": _txt(sh.raw(r, 7)),
                "weights": "auto",
            })
        elif block == "transport":
            transports.append({
                "name": a,
                "transport_id": _txt(sh.raw(r, 2)),
                "supplier_names": [x.strip() for x in _txt(sh.raw(r, 3)).split(",") if x.strip()],
            })
    return buyers, suppliers, transports


# ------------------------------------------------------------ item layouts
def parse_pp(sh: Sheet, supplier_code: str, group: str):
    """Kiran · Hansa-PP · VP-PP.

    A code · B gd · C gl · D size · E length · F pack/unit · G pack/box
    H description · I barcode · J hsn · O net/box · P gross/box
    T bg/box · U p/box · X type-up · AA purchase ₹/pc · AD FOB $/100pcs
    Last row is the TOTAL line.
    """
    out = []
    for r in range(3, sh.max_row + 1):
        code = _txt(sh.raw(r, 1))
        if not code or _txt(sh.raw(r, 9)).upper() == "TOTAL":
            continue
        spec = sticker_spec(sh, r, ttl_col=22, labels_col=23, sheets_col=25, up_col=24)
        out.append({
            "code": code,
            "gd": _txt(sh.raw(r, 2)) or code,
            "oswin": "",
            "gl": _txt(sh.raw(r, 3)),
            "size": _txt(sh.raw(r, 4)),
            "length": _txt(sh.raw(r, 5)),
            "pack_unit": int(_numf(sh.raw(r, 6))),
            "packing": int(_numf(sh.raw(r, 7))),
            "description": _txt(sh.raw(r, 8)),
            "barcode": _txt(sh.raw(r, 9)),
            "hsn": hsn8(sh.raw(r, 10)),
            "volume": FLAT_VOLUME,
            "net_per_box": _numf(sh.val(r, 15)),
            "gross_per_box": _numf(sh.val(r, 16)),
            "bg_per_box": _numf(sh.val(r, 20)),
            "p_per_box": _numf(sh.val(r, 21)),
            **spec,
            "sticker_rule": "pp",
            "uom": "PCS",
            "value_mode": "piece",
            "unit_value": _numf(sh.raw(r, 27)),
            "fob_mode": "100",
            "unit_fob100": _numf(sh.raw(r, 30)),
            "group": group,
            "supplier_code": supplier_code,
            "source_sheet": sh.name,
            "source_row": r,
        })
    return out


def parse_grn(sh: Sheet, supplier_code: str, group: str):
    """Hansa-GRN · VP-GRN.

    A gd · B gl · C size · D pack/unit · E pack/box · F description
    G barcode · H hsn · M net/box · N gross/box · R bg/box · S p/box
    V type-up · Y purchase ₹/pc · AB FOB $/100pcs
    """
    out = []
    for r in range(3, sh.max_row + 1):
        gd = _txt(sh.raw(r, 1))
        if not gd or _txt(sh.raw(r, 7)).upper() == "TOTAL":
            continue
        spec = sticker_spec(sh, r, ttl_col=20, labels_col=21, sheets_col=23, up_col=22)
        out.append({
            "code": gd,
            "gd": gd,
            "oswin": "",
            "gl": _txt(sh.raw(r, 2)),
            "size": _txt(sh.raw(r, 3)),
            "length": "",
            "pack_unit": int(_numf(sh.raw(r, 4))),
            "packing": int(_numf(sh.raw(r, 5))),
            "description": _txt(sh.raw(r, 6)),
            "barcode": _txt(sh.raw(r, 7)),
            "hsn": hsn8(sh.raw(r, 8)),
            "volume": FLAT_VOLUME,
            "net_per_box": _numf(sh.val(r, 13)),
            "gross_per_box": _numf(sh.val(r, 14)),
            "bg_per_box": _numf(sh.val(r, 18)),
            "p_per_box": _numf(sh.val(r, 19)),
            **spec,
            "sticker_rule": "grn",
            "uom": "PCS",
            "value_mode": "piece",
            "unit_value": _numf(sh.raw(r, 25)),
            "fob_mode": "100",
            "unit_fob100": _numf(sh.raw(r, 28)),
            "group": group,
            "supplier_code": supplier_code,
            "source_sheet": sh.name,
            "source_row": r,
        })
    return out


# Oswin runs several blocks down one sheet. Each entry: (first, last, kind, group).
OSWIN_BLOCKS = [
    (3, 139, "moulded", None),      # group comes from the size banner rows
    (144, 149, "tube", "HYDRO TUBES"),
    (155, 155, "tube", "EZE TUBES"),
    (160, 165, "pipe", "PP PLAIN PIPES"),
]
# Banner rows inside the moulded block that name the size family below them.
OSWIN_BANNERS = {2: '15 MM (1/2")', 16: '20 MM (3/4")', 33: '25 MM (1")',
                 52: '32 MM (1.1/4")', 68: '40 MM (1.1/2")', 85: '50 MM (2")',
                 105: "PP PIPES M/F THREADED"}


def parse_oswin(sh: Sheet, supplier_code: str):
    """Oswin — 7A layout.

    A code · B gd · C oswin code · D gl · E size · F length · G pack/unit
    H pack/box · I description · J barcode · K hsn · O volume/box
    R bg/box · S p/box · X purchase ₹/pc · AA FOB $/pc · V = ROUNDUP(labels/n)

    Tube blocks are ordered in metres (quantity = rolls × metres/roll) and the
    "PP PLAIN PIPES" block carries no prices in the workbook — those import at
    zero and are priced in Setup.
    """
    out = []
    banner = OSWIN_BANNERS[2]
    for first, last, kind, fixed_group in OSWIN_BLOCKS:
        for r in range(first, last + 1):
            if r in OSWIN_BANNERS:
                banner = OSWIN_BANNERS[r]
                continue
            code = _txt(sh.raw(r, 1))
            if not code:
                continue
            desc_col, bar_col = (9, 10) if kind != "pipe" else (10, 0)
            spec = sticker_spec(sh, r, ttl_col=20, labels_col=21, sheets_col=22, up_col=0)
            if not spec["type_up"]:
                spec["type_up"] = 125
            out.append({
                "code": code,
                "gd": _txt(sh.raw(r, 2)) or code,
                "oswin": _txt(sh.raw(r, 3)),
                "gl": _txt(sh.raw(r, 4)),
                "size": _txt(sh.raw(r, 5)),
                "length": _txt(sh.raw(r, 6)),
                "pack_unit": int(_numf(sh.raw(r, 7))),
                "packing": int(_numf(sh.raw(r, 8))),
                "description": _txt(sh.raw(r, desc_col)),
                "barcode": _txt(sh.raw(r, bar_col)) if bar_col else "",
                "hsn": hsn8(sh.raw(r, 11)),
                "volume": _numf(sh.raw(r, 15)),
                "net_per_box": 0.0,
                "gross_per_box": 0.0,
                "bg_per_box": _numf(sh.val(r, 18)),
                "p_per_box": _numf(sh.val(r, 19)),
                **spec,
                "sticker_rule": "oswin",
                "uom": "MTR" if kind == "tube" else "PCS",
                "value_mode": "piece",
                "unit_value": _numf(sh.raw(r, 24)),
                "fob_mode": "piece",
                "unit_fob100": _numf(sh.raw(r, 27)),
                "group": fixed_group or banner,
                "supplier_code": supplier_code,
                "source_sheet": sh.name,
                "source_row": r,
            })
    return out


# ------------------------------------------------------------------- main
SHEET_MAP = [
    ("Kiran", parse_pp, "Kiran", "PP Moulded"),
    ("Hansa-PP", parse_pp, "hansa", "PP Moulded"),
    ("Hansa-GRN", parse_grn, "hansa", "GRN Range"),
    ("Oswin", parse_oswin, "Oswin", None),
    ("VP-PP", parse_pp, "VPPlastics", "PP Moulded"),
    ("VP-GRN", parse_grn, "VPPlastics", "GRN Range"),
]


def main():
    wb_f = openpyxl.load_workbook(XLSX, data_only=False)
    wb_v = openpyxl.load_workbook(XLSX, data_only=True)

    buyers, suppliers, transports = parse_parties(Sheet(wb_f, wb_v, "Sheet1"))

    by_name = {s["name"].strip().lower(): s["code"] for s in suppliers}

    def resolve(name: str):
        n = name.strip().lower()
        for full, code in by_name.items():
            if full.startswith(n) or n.startswith(full[:12]):
                return code
        return None

    for t in transports:
        t["supplier_codes"] = [c for c in (resolve(n) for n in t.pop("supplier_names")) if c]

    items = []
    for name, fn, sup, group in SHEET_MAP:
        sh = Sheet(wb_f, wb_v, name)
        rows = fn(sh, sup) if fn is parse_oswin else fn(sh, sup, group)
        items.extend(rows)
        print(f"{name:<10} → {len(rows):>4} items")

    payload = {
        "source": "Docs/Data/Masters.xlsx",
        "buyers": buyers,
        "suppliers": suppliers,
        "transports": transports,
        "items": items,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=1, ensure_ascii=False))
    print(f"\n{len(buyers)} buyers · {len(suppliers)} suppliers · "
          f"{len(transports)} transports · {len(items)} items → {OUT}")


if __name__ == "__main__":
    main()
